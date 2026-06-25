"""
Import/export service for Skills and Talents.

Supported formats:
  - Text (.txt): structured key-value blocks separated by blank lines
  - CSV  (.csv): comma-separated with header row
  - Excel (.xlsx): first worksheet, first row as headers
"""

import csv
import io
import re

# ---------------------------------------------------------------------------
# Text-format parsers
# ---------------------------------------------------------------------------

_BLANK_LINE_RE = re.compile(r'\n[ \t]*\n')

# Keys recognized inside a talent block
_TALENT_KEYS = {
    'nombre':       'name_es',
    'name':         'name_es',
    'descripción':  'description',
    'descripcion':  'description',
    'description':  'description',
}

# Keys recognized inside a skill block (longest-first to avoid prefix clash)
_SKILL_KEYS_ORDERED = [
    ('talentos asociados', 'talentos_asociados'),
    ('talentos_asociados', 'talentos_asociados'),
    ('características',    'caracteristicas'),
    ('caracteristicas',    'caracteristicas'),
    ('características',    'caracteristicas'),
    ('descripción',        'description'),
    ('descripcion',        'description'),
    ('description',        'description'),
    ('tipo',               '_tipo'),
    ('type',               '_tipo'),
    ('nombre',             'name_es'),
    ('name',               'name_es'),
]


def _split_blocks(text: str) -> list[str]:
    # Normalize Windows (\r\n) and old Mac (\r) line endings before splitting
    text = text.replace('\r\n', '\n').replace('\r', '\n')
    return [b.strip() for b in _BLANK_LINE_RE.split(text.strip()) if b.strip()]


def parse_talent_text(text: str) -> list[dict]:
    """
    Parse talent text blocks:

    Nombre: ¡A CORRER!
    Descripción: Cuando tu vida está en peligro…
    """
    results = []
    for block in _split_blocks(text):
        talent = {}
        current_field = None
        for raw_line in block.split('\n'):
            line = raw_line.strip()
            if not line:
                continue
            matched = False
            for prefix, field in _TALENT_KEYS.items():
                if line.lower().startswith(prefix + ':'):
                    val = line[len(prefix) + 1:].strip()
                    talent[field] = val
                    current_field = field
                    matched = True
                    break
            if not matched and current_field == 'description':
                talent['description'] = talent.get('description', '') + ' ' + line
        if 'name_es' in talent:
            if 'description' in talent:
                talent['description'] = talent['description'].strip() or None
            results.append(talent)
    return results


def parse_skill_text(text: str) -> list[dict]:
    """
    Parse skill text blocks:

    Nombre: ACTUAR (Varios)
    Tipo: Avanzada.
    Características: Empatía.
    Descripción: Esta habilidad…
    Talentos asociados: Contorsionista, Imitador.
    """
    results = []
    for block in _split_blocks(text):
        skill = {}
        current_field = None
        for raw_line in block.split('\n'):
            line = raw_line.strip()
            if not line:
                continue
            matched = False
            ll = line.lower()
            for prefix, field in _SKILL_KEYS_ORDERED:
                if ll.startswith(prefix + ':'):
                    val = line[len(prefix) + 1:].strip().rstrip('.')
                    skill[field] = val
                    current_field = field
                    matched = True
                    break
            if not matched and current_field == 'description':
                skill['description'] = skill.get('description', '') + ' ' + line

        if 'name_es' not in skill:
            continue

        # Convert _tipo to is_advanced
        tipo = skill.pop('_tipo', '').lower().rstrip('.')
        skill['is_advanced'] = tipo in ('avanzada', 'advanced')

        # Normalise empty/none values
        for field in ('description', 'caracteristicas', 'talentos_asociados'):
            val = (skill.get(field) or '').strip()
            skill[field] = None if val.lower() in ('ninguno', 'none', '—', '-', '') else val

        results.append(skill)
    return results


# ---------------------------------------------------------------------------
# CSV parsers
# ---------------------------------------------------------------------------

def _csv_rows(file_bytes: bytes) -> tuple[list[str], list[dict]]:
    text = file_bytes.decode('utf-8-sig', errors='replace')
    reader = csv.DictReader(io.StringIO(text))
    headers = [h.lower().strip() for h in (reader.fieldnames or [])]
    rows = [{k.lower().strip(): (v or '').strip() for k, v in row.items()} for row in reader]
    return headers, rows


def parse_talent_csv(file_bytes: bytes) -> list[dict]:
    _, rows = _csv_rows(file_bytes)
    results = []
    for row in rows:
        name = row.get('nombre') or row.get('name_es') or row.get('name') or ''
        if not name.strip():
            continue
        desc = row.get('descripción') or row.get('descripcion') or row.get('description') or ''
        results.append({'name_es': name.strip(), 'description': desc.strip() or None})
    return results


def parse_skill_csv(file_bytes: bytes) -> list[dict]:
    _, rows = _csv_rows(file_bytes)
    results = []
    for row in rows:
        name = row.get('nombre') or row.get('name_es') or row.get('name') or ''
        if not name.strip():
            continue
        tipo = (row.get('tipo') or row.get('type') or '').lower().rstrip('.')
        caract = row.get('características') or row.get('caracteristicas') or row.get('characterisics') or ''
        desc = row.get('descripción') or row.get('descripcion') or row.get('description') or ''
        tal_as = row.get('talentos asociados') or row.get('talentos_asociados') or row.get('related talents') or ''
        results.append({
            'name_es':           name.strip(),
            'is_advanced':       tipo in ('avanzada', 'advanced'),
            'caracteristicas':   caract.strip() or None,
            'description':       desc.strip() or None,
            'talentos_asociados': tal_as.strip() or None,
        })
    return results


# ---------------------------------------------------------------------------
# Excel parsers
# ---------------------------------------------------------------------------

def _excel_rows(file_bytes: bytes) -> tuple[list[str], list[dict]]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], []
    headers = [str(h).lower().strip() if h is not None else '' for h in rows[0]]
    result = []
    for row in rows[1:]:
        if not any(row):
            continue
        result.append({headers[i]: str(v).strip() if v is not None else '' for i, v in enumerate(row) if i < len(headers)})
    return headers, result


def parse_talent_xlsx(file_bytes: bytes) -> list[dict]:
    _, rows = _excel_rows(file_bytes)
    results = []
    for row in rows:
        name = row.get('nombre') or row.get('name_es') or row.get('name') or ''
        if not name.strip():
            continue
        desc = row.get('descripción') or row.get('descripcion') or row.get('description') or ''
        results.append({'name_es': name.strip(), 'description': desc.strip() or None})
    return results


def parse_skill_xlsx(file_bytes: bytes) -> list[dict]:
    _, rows = _excel_rows(file_bytes)
    results = []
    for row in rows:
        name = row.get('nombre') or row.get('name_es') or row.get('name') or ''
        if not name.strip():
            continue
        tipo = (row.get('tipo') or row.get('type') or '').lower().rstrip('.')
        caract = row.get('características') or row.get('caracteristicas') or ''
        desc = row.get('descripción') or row.get('descripcion') or row.get('description') or ''
        tal_as = row.get('talentos asociados') or row.get('talentos_asociados') or ''
        results.append({
            'name_es':           name.strip(),
            'is_advanced':       tipo in ('avanzada', 'advanced'),
            'caracteristicas':   caract.strip() or None,
            'description':       desc.strip() or None,
            'talentos_asociados': tal_as.strip() or None,
        })
    return results


# ---------------------------------------------------------------------------
# Format dispatcher
# ---------------------------------------------------------------------------

def parse_talents(file_bytes: bytes, extension: str) -> list[dict]:
    ext = extension.lower().lstrip('.')
    if ext in ('xlsx', 'xls'):
        return parse_talent_xlsx(file_bytes)
    if ext == 'csv':
        return parse_talent_csv(file_bytes)
    # Default: plain text
    return parse_talent_text(file_bytes.decode('utf-8-sig', errors='replace'))


def parse_skills(file_bytes: bytes, extension: str) -> list[dict]:
    ext = extension.lower().lstrip('.')
    if ext in ('xlsx', 'xls'):
        return parse_skill_xlsx(file_bytes)
    if ext == 'csv':
        return parse_skill_csv(file_bytes)
    return parse_skill_text(file_bytes.decode('utf-8-sig', errors='replace'))


# ---------------------------------------------------------------------------
# Exporters — plain text
# ---------------------------------------------------------------------------

def export_talents_text(talents) -> str:
    parts = []
    for t in talents:
        block = [f'Nombre: {t.name_es}']
        if t.description:
            block.append(f'Descripción: {t.description}')
        parts.append('\n'.join(block))
    return '\n\n'.join(parts) + '\n'


def export_skills_text(skills) -> str:
    parts = []
    for s in skills:
        block = [f'Nombre: {s.name_es}']
        block.append(f"Tipo: {'Avanzada' if s.is_advanced else 'Básica'}.")
        if s.caracteristicas:
            block.append(f'Características: {s.caracteristicas}.')
        if s.description:
            block.append(f'Descripción: {s.description}')
        block.append(f"Talentos asociados: {s.talentos_asociados or 'Ninguno'}.")
        parts.append('\n'.join(block))
    return '\n\n'.join(parts) + '\n'


# ---------------------------------------------------------------------------
# Exporters — CSV
# ---------------------------------------------------------------------------

def export_talents_csv(talents) -> bytes:
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(['nombre', 'descripcion'])
    for t in talents:
        w.writerow([t.name_es, t.description or ''])
    return out.getvalue().encode('utf-8-sig')


def export_skills_csv(skills) -> bytes:
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(['nombre', 'tipo', 'caracteristicas', 'descripcion', 'talentos_asociados'])
    for s in skills:
        w.writerow([
            s.name_es,
            'Avanzada' if s.is_advanced else 'Básica',
            s.caracteristicas or '',
            s.description or '',
            s.talentos_asociados or '',
        ])
    return out.getvalue().encode('utf-8-sig')


# ---------------------------------------------------------------------------
# Exporters — Excel
# ---------------------------------------------------------------------------

def export_talents_xlsx(talents) -> bytes:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Talentos'
    ws.append(['nombre', 'descripcion'])
    for t in talents:
        ws.append([t.name_es, t.description or ''])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_skills_xlsx(skills) -> bytes:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Habilidades'
    ws.append(['nombre', 'tipo', 'caracteristicas', 'descripcion', 'talentos_asociados'])
    for s in skills:
        ws.append([
            s.name_es,
            'Avanzada' if s.is_advanced else 'Básica',
            s.caracteristicas or '',
            s.description or '',
            s.talentos_asociados or '',
        ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
